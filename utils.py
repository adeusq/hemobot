import openpyxl
import pyperclip
import pyautogui
from tkinter import Tk, messagebox, filedialog
import tkinter as tk
import time
import pandas as pd
import re
import tempfile
import os

def criar_arquivo_modelo():

    caminho_arquivo_modelo = filedialog.asksaveasfilename(
        title="Salvar Arquivo Modelo",
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
    )

    if caminho_arquivo_modelo:
        workbook = openpyxl.Workbook()
        pagina = workbook.active
        pagina.title = "Extraction-Log"

    
    headers = (['PF', 'Amostra', 'ABO', 'RhD', 'Fenotipagem'])
    pagina.append(headers)
    
    for _ in range(20): 
        pagina.append([None] * len(headers))
    
    try:
        workbook.save(caminho_arquivo_modelo)
        messagebox.showinfo("Sucesso", f"Arquivo modelo criado com sucesso em {caminho_arquivo_modelo}.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar o arquivo modelo: {e}")

def preencher_planilha(linha_inicio):
    root = Tk()
    root.withdraw()

    arquivo_caminho = filedialog.asksaveasfilename(
    title="Selecione o arquivo Excel para automatizar",
    defaultextension=".xlsx",
    filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
    )

    if not arquivo_caminho:
        messagebox.showwarning("Nenhum arquivo selecionado", "Por favor, selecione um arquivo para continuar.")
        return
    
    
    workbook = openpyxl.load_workbook(arquivo_caminho)
    pagina_genotipagem = workbook['Extraction-Log']
    
    indice_coluna_destino = 0
    indice_coluna_abo = 2
    indice_coluna_rhd = 3
    indice_coluna_fenotipagem = 4
    
    linhas_para_atualizar = [
        linha for linha in pagina_genotipagem.iter_rows(min_row=linha_inicio)
        if linha[indice_coluna_destino].value is None
    ]

    if not linhas_para_atualizar:
        messagebox.showinfo("Nenhuma atualização", "Não há linhas para atualizar a partir da linha especificada.")
        return

    for linha in pagina_genotipagem.iter_rows(min_row=linha_inicio):
        if linha[indice_coluna_destino].value is not None:
            continue

        num_amostra = linha[1].value
        pyperclip.copy(num_amostra)
        
        # Captura da PF
        pyautogui.click(223,361)
        pyautogui.write('=')
        pyautogui.hotkey('ctrl', 'v')
        pyautogui.write('006')
        pyautogui.press('enter')
        pyautogui.sleep(2)
        pyautogui.click(214,381)  
        pyautogui.click(214,381) 
        pyautogui.hotkey('ctrl', 'c') 
        pyautogui.hotkey('ctrl', 'v')        
        info_campo = pyperclip.paste()
        linha[indice_coluna_destino].value = info_campo 
                
        # Captura Tipagem ABO
        pyautogui.click(381, 385)
        pyautogui.click(381, 385) 
        pyautogui.hotkey('ctrl', 'c') 
        pyautogui.hotkey('ctrl', 'v')       
        abo = pyperclip.paste()
        linha[indice_coluna_abo].value = abo

        # Captura Tipagem RhD
        pyautogui.click(470, 386)  
        pyautogui.click(470, 386)  
        pyautogui.hotkey('ctrl', 'c') 
        pyautogui.hotkey('ctrl', 'v')      
        rhd = pyperclip.paste()
        linha[indice_coluna_rhd].value = rhd
        
    try:
        workbook.save(arquivo_caminho)
        messagebox.showinfo("Arquivo salvo", f"Arquivo atualizado e salvo com sucesso em {arquivo_caminho}.")
    except Exception as e:
        messagebox.showerror("Erro ao salvar", f"Ocorreu um erro ao salvar o arquivo: {e}")

def atualizar_progresso(progresso, valor):
    progresso['value'] = valor
    progresso.update_idletasks()

def export_columns_to_txt(txt_file_input, origem_file, origem_sheet, txt_file_output, update_progress=None):
    codigo_column_index_origem = 8  # Índice da coluna do código de extração no arquivo de origem
    amostra_column_index_origem = 1  # Índice da coluna onde está o número da amostra no arquivo de origem

    workbook_origem = openpyxl.load_workbook(origem_file)
    sheet_origem = workbook_origem[origem_sheet]

    codigo_para_amostra = {}
    for row in sheet_origem.iter_rows(min_row=2):
        codigo = str(row[codigo_column_index_origem].value).strip()
        amostra = str(row[amostra_column_index_origem].value).strip()

        if codigo and amostra:
            codigo_para_amostra[codigo] = amostra

    if not codigo_para_amostra:
        messagebox.showwarning("Aviso", "Nenhum dado correspondente encontrado no arquivo de origem.")
        return

    with open(txt_file_input, 'r', encoding='utf-8') as infile:
        linhas = infile.readlines()
        total_linhas = len(linhas)
        
        with open(txt_file_output, 'w', encoding='utf-8') as outfile:
            for i, linha in enumerate(linhas):
                codigo = linha.strip()
                amostra = codigo_para_amostra.get(codigo, "Amostra não encontrada")

                outfile.write(f"{amostra}\t{codigo}\n")

                if update_progress:
                    progress = (i + 1) / total_linhas * 100
                    update_progress(progress)
                    time.sleep(0.05)  
                
    messagebox.showinfo("Sucesso", f'Dados exportados para {txt_file_input} com sucesso!')
   
def clean_column_name(col_name):
    return re.sub(r'\s*\(.*?\)\s*', '', col_name)

def converter_xls_para_xlsx(xls_file_path):
    try:
        input_directory = os.path.dirname(xls_file_path)
        
        temp_xlsx_file_path = os.path.join(input_directory, os.path.basename(xls_file_path).replace('.xls', '.xlsx'))

        xls = pd.ExcelFile(xls_file_path)

        with pd.ExcelWriter(temp_xlsx_file_path, engine='openpyxl') as writer:
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                
                if sheet_name == 'ID CORE XT Fenótipo':
                    df = df.drop(index=range(0, 19))
                    df = df.head(50)
                    df = df.reset_index(drop=True)
                
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        return temp_xlsx_file_path

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao converter o arquivo: {e}")
        return None

def processar_fenotipagem(file_path, output_directory):
    try:
        df = pd.read_excel(file_path, sheet_name='ID CORE XT Fenótipo')

        resultados = []

        for index, row in df.iterrows():
            amostra = row.iloc[0]
            match = re.search(r'B315\d+|B3121\d+', amostra)
            if match:
                amostra_id = match.group()
            else:
                amostra_id = amostra

            antigenos = []
            observacoes = set()
            sangue_raro_msg = []

            valor_s = None
            valor_s_pequeno = None
            valor_m = valor_n = None
            valor_jka = valor_jkb = None
            valor_hrb = valor_e = None

            for col in df.columns[1:]:
                value = row[col]
                if isinstance(value, float) and value == 0.0:
                    value = 0
                if not pd.isna(value) and (value in ['+', '0', 'NC', 'UN'] or isinstance(value, str) and re.match(r'\+\(\d+\)', value)) or value == 0:
                    col_name = clean_column_name(col)
                    antigenos.append(f"{col_name}({value})")
                    
                    if re.search(r'\(2\)', str(value)):
                        observacoes.add("(2) Expressão débil de antígenos como detectado por alguns anticorpos.")
                    if re.search(r'\(3\)', str(value)):
                        observacoes.add("(3) Expressão parcial de antígenos como detectado por alguns anticorpos.")
                    if re.search(r'\(6\)', str(value)):
                        observacoes.add("(6) Expressão parcial de antígenos como detectado por alguns anticorpos.")
                    if re.search(r'\(7\)', str(value)):
                        observacoes.add("(7) Expressão parcial ou fraca de antígenos como detectado por alguns anticorpos.")
                    if re.search(r'\(29\)', str(value)):
                        observacoes.add("(29) Expressão parcial de antígenos como detectado por alguns anticorpos.")

                    if col_name == 'Kpb' and value == 0:
                        sangue_raro_msg.append('Kpb-')
                    if col_name == 'k' and value == 0:
                        sangue_raro_msg.append('k-')

                    if col_name == 'S':
                        valor_s = value

                    if col_name == 's':
                        valor_s_pequeno = value

                    if col_name == 'M':
                        valor_m = value  
                    if col_name == 'N':
                        valor_n = value 

                    if col_name == 'Jka':
                        valor_jka = value 
                    if col_name == 'Jkb':
                        valor_jkb = value 

                    if col_name == 'hrB':
                        valor_hrb = value
                    if col_name == 'e':
                        valor_e = value

                    if col_name == 'Dib' and value == 0:
                        sangue_raro_msg.append('Dib-')
                    if col_name == 'Lub' and value == 0:
                        sangue_raro_msg.append('Lub-')
                    if col_name == 'Coa' and value == 0:
                        sangue_raro_msg.append('Coa-')
                    if col_name == 'Yta' and value == 0:
                        sangue_raro_msg.append('Yta-')
                    if col_name == 'Jsb' and value == 0:
                        sangue_raro_msg.append('Jsb-')
                    if col_name == 'Hy' and value == 0:
                        sangue_raro_msg.append('Hy-')
                    if col_name == 'Joa' and value == 0:
                        sangue_raro_msg.append('Joa-')

            # Verificar combinação S(0) e s(0)
            try:
                valor_s = float(valor_s)
            except (ValueError, TypeError):
                valor_s = None

            try:
                valor_s_pequeno = float(valor_s_pequeno)
            except (ValueError, TypeError):
                valor_s_pequeno = None

            if valor_s == 0 and valor_s_pequeno == 0:
                sangue_raro_msg.append('S- e s-')

            try:
                valor_n = float(valor_n)
            except (ValueError, TypeError):
                valor_n = None

            if valor_m == 0 and valor_n == 0:
                sangue_raro_msg.append('M- e N-')

            # Verificar combinação Jka(0) e Jkb(0)
            try:
                valor_jka = float(valor_jka)
            except (ValueError, TypeError):
                valor_jka = None

            try:
                valor_jkb = float(valor_jkb)
            except (ValueError, TypeError):
                valor_jkb = None

            if valor_jka == 0 and valor_jkb == 0:
                sangue_raro_msg.append('Jka- e Jkb-')

            # Verificar combinação hrB(0)
            try:
                valor_hrb = float(valor_hrb)
            except (ValueError, TypeError):
                valor_hrb = None

            if isinstance(valor_e, str) and '+' in valor_e:
                valor_e = '+'
            else:
                valor_e = None

            if valor_hrb == 0 and valor_e == '+':
                sangue_raro_msg.append('hrB-')  

            categories = [
                (0, 9), (9, 15), (15, 17), (17, 19), (19, 25),
                (25, 27), (27, 31), (31, 33), (33, 35), (35, None)
            ]

            antigenos_str = '\n'.join([
                ', '.join(antigenos[start:end] if end is not None else antigenos[start:])
                for start, end in categories
            ])

            resultado = f"{amostra_id}:\nFenotipagem deduzida a partir da genotipagem;\n{antigenos_str}".rstrip('; ').rstrip('.')
            
            if observacoes:
                resultado += "\n\n" + "\n".join(observacoes)

            if sangue_raro_msg:
                resultado += f"\n\nIndivíduo com SANGUE RARO [{', '.join(sangue_raro_msg)}]. Caso necessite transfundir, entrar em contato com o hemocentro coordenador."

            resultados.append(resultado)
            
        input_filename = os.path.splitext(os.path.basename(file_path))[0]
        output_file_path = os.path.join(output_directory, f"Resultado {input_filename}.txt")

        with open(output_file_path, 'w', encoding='utf-8') as f:
            for resultado in resultados:
                f.write(resultado + '\n\n')

        messagebox.showinfo("Sucesso", f"Dados salvos com sucesso em: {os.path.abspath(output_file_path)}")
    
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao processar o arquivo: {e}")
